from __future__ import annotations
from abc import ABC, abstractmethod
from typing import Any, List, Set
import openpyxl as op
from openpyxl.worksheet.worksheet import Worksheet
from excel_checker.visitorum.visitor import Visitor
from sortedcollections import SortedSet

class Component(ABC):
    """
    The Component interface declares an `accept` method that should take the
    base visitor interface as an argument.
    """

    @abstractmethod
    def accept(self, visitor: Visitor) -> None:
        pass


class CWorkbook(Component):
    def __init__(self, wbk: op.Workbook):
        self.wbk = wbk
        self.sheets: Set[CWorksheet] = set()

    def accept(self, visitor: Visitor) -> None:
        return visitor.visit_workbook(self)

    def getTitle(self) -> str:
        return "Title still to give"

    def addSheet(self, name: str) -> CWorksheet:
        if name in self.wbk.sheetnames:
            ws = CWorksheet(self.wbk[name])
            self.sheets.add(ws)
            return ws
        else:
            print(f"Sheet {name} not present")

    def addAllSheets(self, name: str) -> List[CWorksheet]:
        for name in self.wbk.sheetnames:
            self.addSheet(name)
        return self.sheets

    def __hash__(self):
        return hash((self.wbk))

    def __eq__(self, other):
        if not isinstance(other, type(self)): return NotImplemented
        return self.wbk == other.wbk


class CWorksheet(Component):
    def __init__(self, wsh: Worksheet):
        self.wsh = wsh
        self.rows: SortedSet(CRow) = SortedSet()
        self.columns: SortedSet(CCol) = SortedSet()
        self.cells: SortedSet(CCell) = SortedSet()    

    def accept(self, visitor: Visitor) -> None:
        return visitor.visit_worksheet(self)

    def getTitle(self) -> str:
        return self.wsh.title

    def addRow(self, row: int) -> CRow:
        r = CRow(self.wsh, row)
        self.rows.add(r)
        return r

    def addCol(self, col: int) -> CCol:
        c = CCol(self.wsh, col)
        self.columns.add(c)
        return c

    def addCell(self, row: int, col: int) -> CCell:
        c = CCell(self.wsh, row, col)
        self.cells.add(c)
        return c
    
    def __hash__(self):
        return hash((self.wsh))

    def __eq__(self, other):
        if not isinstance(other, type(self)): return NotImplemented
        return self.wsh == other.wsh

class CRow(Component):
    def __init__(self, sheet: Worksheet, rownum: int):
        self.i = rownum
        self.sheet = sheet

    def accept(self, visitor: Visitor) -> None:
        return visitor.visit_row(self)

    def getRowNum(self) -> str:
        return self.i

    def getColInRow(self, col: int) -> Any:
        return self.sheet.cell(self.i, col).value
    
    def __hash__(self):
        return hash((self.sheet, self.i))

    def __eq__(self, other):
        if not isinstance(other, type(self)): return NotImplemented
        return self.sheet == other.sheet and self.i == other.i
    
    def __lt__(self, other):
        if not isinstance(other, type(self)): return NotImplemented
        return self.i < other.i

    def __str__(self):
        return f"{self.i}"

    def __repr__(self):
        return f"{self.i}"


class CCol(Component):
    def __init__(self, sheet: Worksheet, colnum: int):
        self.j = colnum
        self.sheet = sheet

    def accept(self, visitor: Visitor) -> None:
        return visitor.visit_column(self)

    def getColNum(self) -> str:
        return self.j

    def getRowInCol(self, row: int) -> Any:
        return self.sheet.cell(row, self.j).value

    def __hash__(self):
        return hash((self.sheet, self.j))

    def __eq__(self, other):
        if not isinstance(other, type(self)): return NotImplemented
        return self.sheet == other.sheet and self.j == other.j

    def __lt__(self, other):
        if not isinstance(other, type(self)): return NotImplemented
        return self.j < other.j

    def __str__(self):
        return f"{self.j}"

    def __repr__(self):
        return f"{self.j}"

class CCell(Component):
    def __init__(self, sheet: Worksheet, rownum: int, colnum: int, alias=""):
        self.i = rownum
        self.j = colnum
        self.alias = alias
        self.sheet = sheet

    def accept(self, visitor: Visitor) -> None:
        return visitor.visit_cell(self)

    def getPosition(self) -> str:
        return (self.i, self.j)

    def getAlias(self) -> str:
        return self.alias

    def getValue(self) -> Any:
        return self.sheet.cell(self.i, self.j).value
    
    def __hash__(self):
        return hash((self.sheet, self.i, self.j))

    def __eq__(self, other):
        if not isinstance(other, type(self)): return NotImplemented
        return self.sheet == other.sheet and self.i == other.i and self.j == other.j

    def __lt__(self, other):
        if not isinstance(other, type(self)): return NotImplemented
        if self.i < other.i:
            return True
        elif self.i == other.i: 
            if self.j < other.j:
                return True
            return False
        else:
            return False

    def __str__(self):
        return f"({self.i},{self.j})"

    def __repr__(self):
        return f"({self.i},{self.j})"
