from __future__ import annotations
from abc import ABC, abstractmethod
from typing import Any, List
import openpyxl as op
from openpyxl.worksheet.worksheet import Worksheet
from visitorum.visitor import Visitor


class Component(ABC):
    """
    The Component interface declares an `accept` method that should take the
    base visitor interface as an argument.
    """

    @abstractmethod
    def accept(self, visitor: Visitor) -> None:
        pass
    
class CWorkbook(Component):
    
    def __init__(self, wbk: op.Workbook):
        self.wbk = wbk
        self.sheets: List[CWorksheet] = []

    def accept(self, visitor: Visitor) -> None:
        return visitor.visit_workbook(self)

    def getTitle(self) -> str:
        return "Title still to give"
    
    def addSheet(self, name: str) -> CWorksheet:
        if name in self.wbk.sheetnames:
            ws = CWorksheet(self.wbk[name])
            self.sheets.append(ws)
            return ws 
        else:
            print(f"Sheet {name} not present")

    def addAllSheets(self, name: str) -> List[CWorksheet]:
        for name in self.wbk.sheetnames:
            self.sheets.append(CWorksheet(self.wbk[name]))

class CWorksheet(Component):
    
    def __init__(self, wsh: Worksheet):
        self.wsh = wsh
        self.rows : List[CRow] = []
        self.columns : List[CCol] = []
        self.cells : List[CCell] = []

    def accept(self, visitor: Visitor) -> None:
        return visitor.visit_worksheet(self)

    def getTitle(self) -> str:
        return self.wsh.title

    def addRow(self, row: int) -> CRow:
        r = CRow(self.wsh, row)
        self.rows.append(r)
        return r

    def addCol(self, col: int) -> CCol:
       c = CCol(self.wsh, col)
       self.columns.append(c)
       return c
    
    def addCell(self, row: int, col: int) -> CCell:
       c = CCell(self.wsh, row, col)
       self.cells.append(c)
       return c

class CRow(Component):
    
    def __init__(self, sheet: Worksheet, rownum : int):
        self.i = rownum
        self.sheet = sheet

    def accept(self, visitor: Visitor) -> None:
        return visitor.visit_row(self)

    def getRowNum(self) -> str:
        return self.i
    
    def getColInRow(self, col:int) -> Any:
        return self.sheet.cell(self.i, col).value

class CCol(Component):
    
    def __init__(self, sheet: Worksheet, colnum : int):
        self.j = colnum
        self.sheet = sheet

    def accept(self, visitor: Visitor) -> None:
        return visitor.visit_column(self)

    def getColNum(self) -> str:
        return self.j
    
    def getRowInCol(self, row:int) -> Any:
        return self.sheet.cell(row, self.j).value


class CCell(Component):
    
    def __init__(self, sheet: Worksheet, rownum : int, colnum : int, alias=""):
        self.i = rownum
        self.j = colnum
        self.alias = alias
        self.sheet = sheet

    def accept(self, visitor: Visitor) -> None:
        return visitor.visit_cell(self)

    def getPosition(self) -> str:
        return (self.i, self.j)
    
    def getAlias(self) -> str:
        return self.alias

    def getValue(self) -> Any:
        return self.sheet.cell(self.i, self.j).value
